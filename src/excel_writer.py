"""
excel_writer.py
---------------
Handles writing DataFrames to Excel files with support for incremental updates,
custom reports, and end-of-day summaries.

Author: Jonathan Wardwell, Copilot, GPT-4o
License: MIT
"""

import os
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional
from openpyxl import load_workbook, Workbook
from logger import LoggerFactory


class ExcelWriter:
    """
    ExcelWriter handles various Excel output scenarios:
    - Daily batch reports (legacy)
    - Incremental hourly updates
    - On-demand custom reports
    - End-of-day summary reports
    """
    
    def __init__(self, output_dir: str = 'data/output', log_file: str = 'report.log'):
        self.output_dir = output_dir
        self.logger = LoggerFactory.get_logger('excel_writer', log_file)
        # Don't create directory in __init__ - let each method handle it

    def write(self, report_data: dict, date_str: str) -> str:
        """
        Legacy method: Write the report_data to an Excel file named report_<date_str>.xlsx.
        
        Args:
            report_data: { sheet_name: [DataFrame, ...] }
            date_str: Date string for filename
            
        Returns:
            Path to created file
        """
        if not report_data:
            self.logger.warning('No data to write to Excel.')
            return None

        output_path = os.path.join(self.output_dir, f'report_{date_str}.xlsx')
        self.logger.info(f'Writing Excel report to {output_path}')

        try:
            # Ensure output directory exists
            os.makedirs(self.output_dir, exist_ok=True)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df_list in report_data.items():
                    combined_df = pd.concat(df_list, ignore_index=True)
                    combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
            self.logger.info('Excel report generation complete.')
            return output_path
        except Exception as e:
            self.logger.error(f'Failed to write Excel report: {e}', exc_info=True)
            return None

    def write_incremental(self, report_data: Dict, date_str: str, hour_str: str) -> str:
        """
        Write incremental hourly data to existing daily Excel file or create new one.
        
        Args:
            report_data: { sheet_name: [DataFrame, ...] }
            date_str: Date string (YYYY-MM-DD)
            hour_str: Hour string (YYYY-MM-DD_HH)
            
        Returns:
            Path to updated file
        """
        if not report_data:
            self.logger.info('No incremental data to write.')
            return None

        daily_file_path = os.path.join(self.output_dir, f'daily_{date_str}.xlsx')
        self.logger.info(f'Writing incremental data for {hour_str} to {daily_file_path}')

        try:
            # Ensure output directory exists
            os.makedirs(self.output_dir, exist_ok=True)
            
            # Load existing workbook or create new one
            if os.path.exists(daily_file_path):
                workbook = load_workbook(daily_file_path)
                existing_data = self._load_existing_data(daily_file_path)
            else:
                workbook = Workbook()
                # Remove default sheet
                if 'Sheet' in workbook.sheetnames:
                    workbook.remove(workbook['Sheet'])
                existing_data = {}

            # Merge new data with existing
            merged_data = self._merge_data(existing_data, report_data)

            # Write merged data back to workbook
            self._write_to_workbook(workbook, merged_data)
            
            # Save workbook
            workbook.save(daily_file_path)
            workbook.close()
            
            self.logger.info(f'Incremental update completed for {hour_str}')
            return daily_file_path

        except Exception as e:
            self.logger.error(f'Failed to write incremental data: {e}', exc_info=True)
            return None

    def write_custom(self, report_data: Dict, filename: str) -> str:
        """
        Write custom report with specified filename.
        
        Args:
            report_data: { sheet_name: [DataFrame, ...] }
            filename: Custom filename (should include .xlsx extension)
            
        Returns:
            Path to created file
        """
        if not report_data:
            self.logger.warning('No data for custom report.')
            return None

        output_path = os.path.join(self.output_dir, filename)
        self.logger.info(f'Writing custom report to {output_path}')

        try:
            # Ensure output directory exists
            os.makedirs(self.output_dir, exist_ok=True)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df_list in report_data.items():
                    combined_df = pd.concat(df_list, ignore_index=True)
                    combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            self.logger.info(f'Custom report generated: {filename}')
            return output_path

        except Exception as e:
            self.logger.error(f'Failed to write custom report: {e}', exc_info=True)
            return None

    def write_summary(self, enhanced_data: Dict, filename: str) -> str:
        """
        Write comprehensive end-of-day summary report with enhanced formatting.
        
        Args:
            enhanced_data: { sheet_name: [DataFrame, ...] } including metadata
            filename: Summary filename (should include .xlsx extension)
            
        Returns:
            Path to created file
        """
        if not enhanced_data:
            self.logger.warning('No data for summary report.')
            return None

        output_path = os.path.join(self.output_dir, filename)
        self.logger.info(f'Writing summary report to {output_path}')

        try:
            # Ensure output directory exists
            os.makedirs(self.output_dir, exist_ok=True)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write metadata sheet first if it exists
                if 'Report_Metadata' in enhanced_data:
                    metadata_df = enhanced_data['Report_Metadata'][0]
                    metadata_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Write all data sheets
                for sheet_name, df_list in enhanced_data.items():
                    if sheet_name == 'Report_Metadata':
                        continue  # Already handled
                    
                    combined_df = pd.concat(df_list, ignore_index=True)
                    
                    # Add summary statistics for each sheet
                    summary_stats = self._calculate_sheet_summary(combined_df, sheet_name)
                    
                    # Write data
                    combined_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=5)
                    
                    # Write summary stats at top
                    summary_stats.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

            self.logger.info(f'Summary report generated: {filename}')
            return output_path

        except Exception as e:
            self.logger.error(f'Failed to write summary report: {e}', exc_info=True)
            return None

    def _load_existing_data(self, file_path: str) -> Dict:
        """Load existing data from Excel file."""
        existing_data = {}
        
        try:
            excel_file = pd.ExcelFile(file_path, engine='openpyxl')
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                existing_data[sheet_name] = [df]
            excel_file.close()
            
        except Exception as e:
            self.logger.warning(f'Could not load existing data from {file_path}: {e}')
            
        return existing_data

    def _merge_data(self, existing_data: Dict, new_data: Dict) -> Dict:
        """Merge new data with existing data."""
        merged = existing_data.copy()
        
        for sheet_name, df_list in new_data.items():
            if sheet_name in merged:
                merged[sheet_name].extend(df_list)
            else:
                merged[sheet_name] = df_list
                
        return merged

    def _write_to_workbook(self, workbook, data: Dict):
        """Write data dictionary to openpyxl workbook."""
        # Clear existing sheets
        for sheet_name in list(workbook.sheetnames):
            workbook.remove(workbook[sheet_name])
        
        # Add data sheets
        for sheet_name, df_list in data.items():
            combined_df = pd.concat(df_list, ignore_index=True)
            
            # Create worksheet
            worksheet = workbook.create_sheet(title=sheet_name)
            
            # Write headers
            for col_num, column_title in enumerate(combined_df.columns, 1):
                worksheet.cell(row=1, column=col_num, value=column_title)
            
            # Write data
            for row_num, row_data in enumerate(combined_df.values, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col_num, value=cell_value)

    def _calculate_sheet_summary(self, df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        """Calculate summary statistics for a sheet."""
        summary_data = {
            'Metric': [
                'Total Records',
                'Data Columns',
                'Time Range (if applicable)',
                'Last Updated'
            ],
            'Value': [
                len(df),
                len(df.columns),
                self._get_time_range(df),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        
        return pd.DataFrame(summary_data)

    def _get_time_range(self, df: pd.DataFrame) -> str:
        """Extract time range from DataFrame if date columns exist."""
        date_columns = [col for col in df.columns if 'date' in col.lower() or 'time' in col.lower()]
        
        if not date_columns or df.empty:
            return 'N/A'
        
        try:
            date_col = date_columns[0]
            if date_col in df.columns:
                dates = pd.to_datetime(df[date_col], errors='coerce').dropna()
                if not dates.empty:
                    return f"{dates.min().strftime('%Y-%m-%d')} to {dates.max().strftime('%Y-%m-%d')}"
        except Exception:
            pass
            
        return 'N/A'
