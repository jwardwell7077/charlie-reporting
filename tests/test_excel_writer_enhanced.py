"""
test_excel_writer_enhanced.py
----------------------------
Tests for the enhanced ExcelWriter class with incremental updates and custom reports.

Author: Jonathan Wardwell, Copilot, GPT-4o
License: MIT
"""

import os
import pandas as pd
import pytest
from datetime import datetime
from unittest.mock import Mock, patch
from openpyxl import load_workbook

# Import the classes we're testing
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from services.report_generator.excel_generator import ExcelWriter


# --- Fixtures ---
@pytest.fixture
def temp_output_dir(tmp_path):
    """Create temporary output directory"""
    output_dir = tmp_path / 'output'
    output_dir.mkdir()
    return str(output_dir)


@pytest.fixture
def sample_data():
    """Create sample transformed data"""
    return {
        'IB_Calls': [pd.DataFrame({
            'email_received_date': ['2025-07-27', '2025-07-27'],
            'email_received_timestamp': ['2025-07-27 14:30', '2025-07-27 15:30'],
            'Agent Name': ['John Doe', 'Jane Smith'],
            'Handle': ['123', '456'],
            'Avg Handle': ['5.5', '6.2']
        })],
        'ACQ': [pd.DataFrame({
            'email_received_date': ['2025-07-27'],
            'email_received_timestamp': ['2025-07-27 14:30'],
            'Agent Name': ['Alice Brown'],
            'Handle': ['789']
        })]
    }


@pytest.fixture
def additional_data():
    """Create additional data for incremental testing"""
    return {
        'IB_Calls': [pd.DataFrame({
            'email_received_date': ['2025-07-27'],
            'email_received_timestamp': ['2025-07-27 16:30'],
            'Agent Name': ['Bob Wilson'],
            'Handle': ['999'],
            'Avg Handle': ['4.8']
        })],
        'Productivity': [pd.DataFrame({
            'email_received_date': ['2025-07-27'],
            'email_received_timestamp': ['2025-07-27 16:30'],
            'Agent Name': ['Carol Davis'],
            'Logged In': ['8:00'],
            'On Queue': ['7:30']
        })]
    }


# --- ExcelWriter Enhanced Tests ---
def test_legacy_write_method(temp_output_dir, sample_data):
    """Test legacy write method still works"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    result_path = writer.write(sample_data, '2025-07-27')
    
    expected_path = os.path.join(temp_output_dir, 'report_2025-07-27.xlsx')
    assert result_path == expected_path
    assert os.path.exists(expected_path)
    
    # Verify content
    df_calls = pd.read_excel(expected_path, sheet_name='IB_Calls')
    assert len(df_calls) == 2
    assert 'Agent Name' in df_calls.columns


def test_write_incremental_new_file(temp_output_dir, sample_data):
    """Test incremental write creating new file"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    result_path = writer.write_incremental(sample_data, '2025-07-27', '2025-07-27_14')
    
    expected_path = os.path.join(temp_output_dir, 'daily_2025-07-27.xlsx')
    assert result_path == expected_path
    assert os.path.exists(expected_path)
    
    # Verify content
    df_calls = pd.read_excel(expected_path, sheet_name='IB_Calls')
    assert len(df_calls) == 2


def test_write_incremental_existing_file(temp_output_dir, sample_data, additional_data):
    """Test incremental write appending to existing file"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    
    # Create initial file
    initial_path = writer.write_incremental(sample_data, '2025-07-27', '2025-07-27_14')
    
    # Add more data
    result_path = writer.write_incremental(additional_data, '2025-07-27', '2025-07-27_16')
    
    assert result_path == initial_path
    
    # Verify merged content
    df_calls = pd.read_excel(result_path, sheet_name='IB_Calls')
    assert len(df_calls) == 3  # 2 original + 1 new
    
    # Check for new sheet
    df_productivity = pd.read_excel(result_path, sheet_name='Productivity')
    assert len(df_productivity) == 1


def test_write_custom_report(temp_output_dir, sample_data):
    """Test custom report generation"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    custom_filename = 'custom_report_test.xlsx'
    
    result_path = writer.write_custom(sample_data, custom_filename)
    
    expected_path = os.path.join(temp_output_dir, custom_filename)
    assert result_path == expected_path
    assert os.path.exists(expected_path)


def test_write_summary_report(temp_output_dir, sample_data):
    """Test summary report with metadata"""
    # Add metadata to sample data
    metadata_df = pd.DataFrame({
        'Metric': ['Report Date', 'Generated At', 'Total Sheets', 'Total Records'],
        'Value': ['2025-07-27', '2025-07-27 18:00:00', '2', '3']
    })
    enhanced_data = sample_data.copy()
    enhanced_data['Report_Metadata'] = [metadata_df]
    
    writer = ExcelWriter(output_dir=temp_output_dir)
    summary_filename = 'eod_summary_2025-07-27.xlsx'
    
    result_path = writer.write_summary(enhanced_data, summary_filename)
    
    expected_path = os.path.join(temp_output_dir, summary_filename)
    assert result_path == expected_path
    assert os.path.exists(expected_path)
    
    # Verify summary sheet exists
    excel_file = pd.ExcelFile(expected_path)
    assert 'Summary' in excel_file.sheet_names
    
    # Check data sheets have summary stats at top
    df_calls = pd.read_excel(expected_path, sheet_name='IB_Calls')
    # Should have summary stats + blank row + headers + data
    assert len(df_calls) >= 6  # 4 summary rows + header + 2 data rows


def test_load_existing_data(temp_output_dir, sample_data):
    """Test loading existing data from Excel file"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    
    # Create initial file
    initial_path = writer.write(sample_data, '2025-07-27')
    
    # Load existing data
    existing_data = writer._load_existing_data(initial_path)
    
    assert 'IB_Calls' in existing_data
    assert 'ACQ' in existing_data
    assert len(existing_data['IB_Calls'][0]) == 2  # 2 rows


def test_merge_data(temp_output_dir, sample_data, additional_data):
    """Test data merging functionality"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    
    merged = writer._merge_data(sample_data, additional_data)
    
    # IB_Calls should have both original and additional data
    assert len(merged['IB_Calls']) == 2  # 2 DataFrames
    
    # Productivity should be new
    assert 'Productivity' in merged
    assert len(merged['Productivity']) == 1
    
    # ACQ should remain unchanged
    assert len(merged['ACQ']) == 1


def test_calculate_sheet_summary(temp_output_dir, sample_data):
    """Test sheet summary calculation"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    
    df = sample_data['IB_Calls'][0]
    summary = writer._calculate_sheet_summary(df, 'IB_Calls')
    
    assert 'Metric' in summary.columns
    assert 'Value' in summary.columns
    assert len(summary) == 4  # 4 summary metrics
    
    # Check specific values
    metrics = summary.set_index('Metric')['Value']
    assert metrics['Total Records'] == 2
    assert metrics['Data Columns'] == 5


def test_get_time_range_with_date_column(temp_output_dir):
    """Test time range extraction with date columns"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    
    df = pd.DataFrame({
        'email_received_date': ['2025-07-27', '2025-07-28'],
        'other_column': ['A', 'B']
    })
    
    time_range = writer._get_time_range(df)
    assert time_range == '2025-07-27 to 2025-07-28'


def test_get_time_range_no_date_column(temp_output_dir):
    """Test time range extraction without date columns"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    
    df = pd.DataFrame({
        'name': ['John', 'Jane'],
        'value': [1, 2]
    })
    
    time_range = writer._get_time_range(df)
    assert time_range == 'N/A'


def test_write_to_workbook(temp_output_dir, sample_data):
    """Test writing data to openpyxl workbook"""
    from openpyxl import Workbook
    
    writer = ExcelWriter(output_dir=temp_output_dir)
    workbook = Workbook()
    
    writer._write_to_workbook(workbook, sample_data)
    
    # Check sheets were created
    assert 'IB_Calls' in workbook.sheetnames
    assert 'ACQ' in workbook.sheetnames
    
    # Check data was written
    calls_sheet = workbook['IB_Calls']
    assert calls_sheet['A1'].value == 'email_received_date'  # Header
    assert calls_sheet['A2'].value == '2025-07-27'  # First data row


def test_empty_data_handling(temp_output_dir):
    """Test handling of empty data"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    
    # Test all methods with empty data
    assert writer.write({}, '2025-07-27') is None
    assert writer.write_incremental({}, '2025-07-27', '2025-07-27_14') is None
    assert writer.write_custom({}, 'custom.xlsx') is None
    assert writer.write_summary({}, 'summary.xlsx') is None


def test_file_creation_error_handling(temp_output_dir):
    """Test error handling when file creation fails"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    
    sample_data = {
        'Test': [pd.DataFrame({'col1': ['value1']})]
    }
    
    # Mock pandas ExcelWriter to raise an exception
    with patch('pandas.ExcelWriter') as mock_excel_writer:
        mock_excel_writer.side_effect = Exception("Simulated write error")
        
        # Should handle errors gracefully
        result = writer.write(sample_data, '2025-07-27')
        assert result is None


def test_incremental_write_error_handling(temp_output_dir, sample_data):
    """Test error handling in incremental write"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    
    # Create a file we can't write to (simulate permission error)
    locked_file = os.path.join(temp_output_dir, 'daily_2025-07-27.xlsx')
    with open(locked_file, 'w') as f:
        f.write('locked')
    
    # Make file read-only (simulate locked file)
    os.chmod(locked_file, 0o444)
    
    try:
        result = writer.write_incremental(sample_data, '2025-07-27', '2025-07-27_14')
        # Should handle error gracefully
        assert result is None
    finally:
        # Restore permissions for cleanup
        os.chmod(locked_file, 0o666)


def test_workbook_operations_integration(temp_output_dir, sample_data, additional_data):
    """Integration test for complete workbook operations"""
    writer = ExcelWriter(output_dir=temp_output_dir)
    
    # Step 1: Create initial incremental file
    path1 = writer.write_incremental(sample_data, '2025-07-27', '2025-07-27_14')
    
    # Step 2: Add more data
    path2 = writer.write_incremental(additional_data, '2025-07-27', '2025-07-27_16')
    
    # Step 3: Generate on-demand report
    ondemand_data = {'IB_Calls': sample_data['IB_Calls']}
    path3 = writer.write_custom(ondemand_data, 'ondemand_test.xlsx')
    
    # Step 4: Generate summary
    metadata_df = pd.DataFrame({
        'Metric': ['Test Metric'],
        'Value': ['Test Value']
    })
    summary_data = sample_data.copy()
    summary_data['Report_Metadata'] = [metadata_df]
    path4 = writer.write_summary(summary_data, 'summary_test.xlsx')
    
    # Verify all files created
    assert path1 == path2  # Incremental should update same file
    assert os.path.exists(path1)
    assert os.path.exists(path3)
    assert os.path.exists(path4)
    
    # Verify incremental file has merged data
    df_calls = pd.read_excel(path1, sheet_name='IB_Calls')
    assert len(df_calls) == 3  # 2 original + 1 additional
    
    # Verify new sheet exists
    assert 'Productivity' in pd.ExcelFile(path1).sheet_names


if __name__ == '__main__':
    pytest.main([__file__])
