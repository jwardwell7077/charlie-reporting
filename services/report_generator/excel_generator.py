"""
Excel Generation Module for Report Generator Service
Migrated and enhanced from src/excel_writer.py
"""
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Any
import asyncio
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

class ExcelGenerator:
    """
    Enhanced Excel generation with async support and advanced formatting.
    Migrated from src/excel_writer.py with improvements for microservices architecture.
    """
    
    def __init__(self, output_dir: str = "data/outputs"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logging.getLogger('excel_generator')
    
    async def generate_report(self, report_data: Dict[str, List[pd.DataFrame]], 
                            filename: str, template: Optional[str] = None) -> str:
        """
        Generate Excel report from processed data.
        
        Args:
            report_data: Dictionary mapping sheet names to DataFrames
            filename: Output filename
            template: Optional template to use
            
        Returns:
            Path to generated Excel file
        """
        if not report_data:
            self.logger.warning('No data provided for Excel generation')
            return None
        
        output_path = self.output_dir / filename
        self.logger.info(f'Generating Excel report: {output_path}')
        
        try:
            # Use asyncio to run the sync Excel generation in a thread
            result = await asyncio.get_event_loop().run_in_executor(
                None, self._generate_excel_sync, report_data, output_path, template
            )
            return str(result)
            
        except Exception as e:
            self.logger.error(f'Failed to generate Excel report: {e}', exc_info=True)
            return None
    
    def _generate_excel_sync(self, report_data: Dict[str, List[pd.DataFrame]], 
                           output_path: Path, template: Optional[str]) -> Path:
        """Synchronous Excel generation (runs in thread pool)."""
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df_list in report_data.items():
                # Combine DataFrames if multiple
                if len(df_list) > 1:
                    combined_df = pd.concat(df_list, ignore_index=True)
                else:
                    combined_df = df_list[0]
                
                # Write to Excel
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply formatting based on template
                if template:
                    self._apply_template_formatting(writer, sheet_name, template)
                else:
                    self._apply_default_formatting(writer, sheet_name)
        
        self.logger.info(f'Excel report generated successfully: {filename}')
        return output_path
    
    async def generate_incremental_report(self, report_data: Dict[str, List[pd.DataFrame]], 
                                        date_str: str, hour_str: str) -> str:
        """
        Generate or update incremental daily report.
        
        Args:
            report_data: New data to add
            date_str: Date string (YYYY-MM-DD)
            hour_str: Hour string (HHMM)
            
        Returns:
            Path to updated Excel file
        """
        daily_filename = f'daily_{date_str}.xlsx'
        daily_file_path = self.output_dir / daily_filename
        
        self.logger.info(f'Updating incremental report for {hour_str}: {daily_file_path}')
        
        try:
            if daily_file_path.exists():
                # Load existing data and append new data
                result = await self._append_to_existing_excel(daily_file_path, report_data, hour_str)
            else:
                # Create new file
                result = await self.generate_report(report_data, daily_filename)
            
            return result
            
        except Exception as e:
            self.logger.error(f'Failed to update incremental report: {e}', exc_info=True)
            return None
    
    async def _append_to_existing_excel(self, file_path: Path, 
                                      new_data: Dict[str, List[pd.DataFrame]], 
                                      hour_str: str) -> str:
        """Append new data to existing Excel file."""
        
        # Read existing data
        existing_data = {}
        with pd.ExcelFile(file_path) as xls:
            for sheet_name in xls.sheet_names:
                existing_data[sheet_name] = pd.read_excel(xls, sheet_name)
        
        # Combine with new data
        combined_data = {}
        for sheet_name, df_list in new_data.items():
            new_df = pd.concat(df_list, ignore_index=True) if len(df_list) > 1 else df_list[0]
            
            if sheet_name in existing_data:
                # Append new data
                combined_df = pd.concat([existing_data[sheet_name], new_df], ignore_index=True)
            else:
                # New sheet
                combined_df = new_df
            
            combined_data[sheet_name] = [combined_df]
        
        # Add any existing sheets not in new data
        for sheet_name, df in existing_data.items():
            if sheet_name not in combined_data:
                combined_data[sheet_name] = [df]
        
        # Generate updated file
        return await self.generate_report(combined_data, file_path.name)
    
    async def generate_summary_report(self, enhanced_data: Dict[str, Any], 
                                    filename: str) -> str:
        """
        Generate comprehensive summary report with metadata.
        
        Args:
            enhanced_data: Data with metadata and summary statistics
            filename: Output filename
            
        Returns:
            Path to generated summary report
        """
        if not enhanced_data:
            self.logger.warning('No data provided for summary report')
            return None
        
        output_path = self.output_dir / filename
        self.logger.info(f'Generating summary report: {output_path}')
        
        try:
            result = await asyncio.get_event_loop().run_in_executor(
                None, self._generate_summary_sync, enhanced_data, output_path
            )
            return str(result)
            
        except Exception as e:
            self.logger.error(f'Failed to generate summary report: {e}', exc_info=True)
            return None
    
    def _generate_summary_sync(self, enhanced_data: Dict[str, Any], output_path: Path) -> Path:
        """Generate summary report synchronously."""
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write main data sheets
            report_data = enhanced_data.get('data', {})
            for sheet_name, df_list in report_data.items():
                combined_df = pd.concat(df_list, ignore_index=True) if len(df_list) > 1 else df_list[0]
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Write summary sheet
            summary_data = enhanced_data.get('summary', {})
            if summary_data:
                summary_df = pd.DataFrame.from_dict(summary_data, orient='index', columns=['Value'])
                summary_df.to_excel(writer, sheet_name='Summary', index=True)
            
            # Write metadata sheet
            metadata = enhanced_data.get('metadata', {})
            if metadata:
                metadata_df = pd.DataFrame.from_dict(metadata, orient='index', columns=['Value'])
                metadata_df.to_excel(writer, sheet_name='Metadata', index=True)
            
            # Apply enhanced formatting
            for sheet_name in writer.sheets:
                self._apply_summary_formatting(writer, sheet_name)
        
        return output_path
    
    def _apply_default_formatting(self, writer, sheet_name: str):
        """Apply default formatting to Excel sheet."""
        worksheet = writer.sheets[sheet_name]
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:  # First row (headers)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_template_formatting(self, writer, sheet_name: str, template: str):
        """Apply template-specific formatting."""
        if template == "executive":
            self._apply_executive_template(writer, sheet_name)
        elif template == "detailed":
            self._apply_detailed_template(writer, sheet_name)
        else:
            self._apply_default_formatting(writer, sheet_name)
    
    def _apply_executive_template(self, writer, sheet_name: str):
        """Apply executive template formatting."""
        worksheet = writer.sheets[sheet_name]
        
        # Executive style: Bold headers, alternating row colors
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        # Format headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Alternating row colors
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            if row_num % 2 == 0:
                for cell in row:
                    cell.fill = light_fill
    
    def _apply_detailed_template(self, writer, sheet_name: str):
        """Apply detailed template formatting."""
        worksheet = writer.sheets[sheet_name]
        
        # Detailed style: Borders, number formatting
        from openpyxl.styles import Border, Side
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to all cells with data
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
        
        self._apply_default_formatting(writer, sheet_name)
    
    def _apply_summary_formatting(self, writer, sheet_name: str):
        """Apply formatting for summary reports."""
        if sheet_name == 'Summary':
            self._format_summary_sheet(writer.sheets[sheet_name])
        elif sheet_name == 'Metadata':
            self._format_metadata_sheet(writer.sheets[sheet_name])
        else:
            self._apply_default_formatting(writer, sheet_name)
    
    def _format_summary_sheet(self, worksheet):
        """Format the summary sheet."""
        # Title formatting
        title_font = Font(bold=True, size=14, color="1F4E79")
        worksheet['A1'].font = title_font
        
        # Key metrics highlighting
        highlight_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        for row in worksheet.iter_rows(min_row=2):
            if any(keyword in str(row[0].value).lower() for keyword in ['total', 'count', 'average']):
                for cell in row:
                    cell.fill = highlight_fill
    
    def _format_metadata_sheet(self, worksheet):
        """Format the metadata sheet."""
        # Metadata styling
        meta_font = Font(italic=True, size=10)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = meta_font
    
    async def get_file_info(self, filename: str) -> Dict[str, Any]:
        """Get information about an existing Excel file."""
        file_path = self.output_dir / filename
        
        if not file_path.exists():
            return {'exists': False, 'error': 'File not found'}
        
        try:
            with pd.ExcelFile(file_path) as xls:
                sheet_info = {}
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name)
                    sheet_info[sheet_name] = {
                        'rows': len(df),
                        'columns': len(df.columns),
                        'column_names': list(df.columns)
                    }
                
                file_stats = file_path.stat()
                return {
                    'exists': True,
                    'size_mb': file_stats.st_size / (1024 * 1024),
                    'modified': datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
                    'sheets': sheet_info,
                    'total_sheets': len(xls.sheet_names)
                }
                
        except Exception as e:
            return {'exists': True, 'error': f'Failed to read file: {str(e)}'}
    
    async def cleanup_old_files(self, days_old: int = 30) -> List[str]:
        """Clean up Excel files older than specified days."""
        cutoff_time = datetime.now().timestamp() - (days_old * 24 * 60 * 60)
        cleaned_files = []
        
        for file_path in self.output_dir.glob('*.xlsx'):
            if file_path.stat().st_mtime < cutoff_time:
                try:
                    file_path.unlink()
                    cleaned_files.append(str(file_path))
                    self.logger.info(f'Cleaned up old file: {file_path}')
                except Exception as e:
                    self.logger.error(f'Failed to clean up {file_path}: {e}')
        
        return cleaned_files
