"""
Test suite for ExcelReportService  
Validates Excel file generation, formatting, and multi-sheet reports
"""
import pytest
import pandas as pd
from pathlib import Path
import sys
import os
from openpyxl import load_workbook

# Add services to path
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..', '..', 'services'))

from report_generator.src.business.services.excel_service import ExcelReportService
from report_generator.src.business.models.report import Report

class TestExcelReportService:
    """Test suite for Excel report generation business logic"""
    
    @pytest.fixture
    def excel_service(self, mock_logger):
        """Create ExcelReportService instance for testing"""
        return ExcelReportService(mock_logger)
    
    @pytest.fixture
    def sample_report_data(self):
        """Sample data for Excel report generation"""
        return {
            "ACQ": pd.DataFrame([
                {"Agent": "John Doe", "Date": "2025-07-28", "Acquisitions": 5, "Revenue": 1000},
                {"Agent": "Jane Smith", "Date": "2025-07-28", "Acquisitions": 3, "Revenue": 750}
            ]),
            "Dials": pd.DataFrame([
                {"Agent": "John Doe", "Date": "2025-07-28", "Dials": 50, "Connects": 15},
                {"Agent": "Jane Smith", "Date": "2025-07-28", "Dials": 45, "Connects": 12}
            ])
        }
    
    def test_excel_file_creation(self, excel_service, sample_report_data, temp_test_dir):
        """Test basic Excel file creation"""
        output_file = temp_test_dir / "test_report.xlsx"
        
        excel_service.create_excel_report(sample_report_data, output_file)
        
        assert output_file.exists()
        assert output_file.stat().st_size > 0
    
    def test_excel_sheet_structure(self, excel_service, sample_report_data, temp_test_dir):
        """Test Excel file has correct sheet structure"""
        output_file = temp_test_dir / "test_report.xlsx"
        
        excel_service.create_excel_report(sample_report_data, output_file)
        
        # Load and verify sheet structure
        workbook = load_workbook(output_file)
        sheet_names = workbook.sheetnames
        
        assert "ACQ" in sheet_names
        assert "Dials" in sheet_names
        assert len(sheet_names) >= 2
    
    def test_excel_data_integrity(self, excel_service, sample_report_data, temp_test_dir):
        """Test Excel file contains correct data"""
        output_file = temp_test_dir / "test_report.xlsx"
        
        excel_service.create_excel_report(sample_report_data, output_file)
        
        # Read back and verify data
        excel_data = pd.read_excel(output_file, sheet_name="ACQ")
        
        assert len(excel_data) == 2
        assert "Agent" in excel_data.columns
        assert "Acquisitions" in excel_data.columns
        assert excel_data["Agent"].iloc[0] == "John Doe"
        assert excel_data["Acquisitions"].iloc[0] == 5
    
    def test_excel_formatting(self, excel_service, sample_report_data, temp_test_dir):
        """Test Excel file formatting and styling"""
        output_file = temp_test_dir / "test_report.xlsx"
        
        excel_service.create_excel_report(sample_report_data, output_file)
        
        workbook = load_workbook(output_file)
        worksheet = workbook["ACQ"]
        
        # Check basic formatting (headers should exist)
        assert worksheet["A1"].value is not None  # Header row exists
        assert worksheet.max_row >= 3  # Headers + 2 data rows
        assert worksheet.max_column >= 4  # All expected columns
    
    def test_multiple_sheet_creation(self, excel_service, sample_report_data, temp_test_dir):
        """Test creation of multiple sheets in Excel file"""
        output_file = temp_test_dir / "multi_sheet_report.xlsx"
        
        excel_service.create_excel_report(sample_report_data, output_file)
        
        workbook = load_workbook(output_file)
        
        # Verify all sheets created
        for sheet_name in sample_report_data.keys():
            assert sheet_name in workbook.sheetnames
            worksheet = workbook[sheet_name]
            assert worksheet.max_row > 1  # Has data beyond headers
    
    def test_empty_data_handling(self, excel_service, temp_test_dir):
        """Test handling of empty data"""
        empty_data = {
            "Empty_Sheet": pd.DataFrame()
        }
        output_file = temp_test_dir / "empty_report.xlsx"
        
        excel_service.create_excel_report(empty_data, output_file)
        
        assert output_file.exists()
        workbook = load_workbook(output_file)
        assert "Empty_Sheet" in workbook.sheetnames
    
    def test_large_dataset_excel(self, excel_service, temp_test_dir, performance_test_data):
        """Test Excel generation with large dataset"""
        import time
        
        large_data = {
            "Large_Dataset": pd.DataFrame(performance_test_data)
        }
        output_file = temp_test_dir / "large_report.xlsx"
        
        start_time = time.time()
        excel_service.create_excel_report(large_data, output_file)
        generation_time = time.time() - start_time
        
        assert output_file.exists()
        assert generation_time < 120  # Should generate within 2 minutes
        
        # Verify data integrity
        workbook = load_workbook(output_file)
        worksheet = workbook["Large_Dataset"]
        assert worksheet.max_row > 1000  # Has substantial data
    
    def test_excel_file_validation(self, excel_service, sample_report_data, temp_test_dir):
        """Test Excel file validation after creation"""
        output_file = temp_test_dir / "validation_test.xlsx"
        
        excel_service.create_excel_report(sample_report_data, output_file)
        
        # Validate file can be opened and read
        try:
            workbook = load_workbook(output_file)
            for sheet_name in sample_report_data.keys():
                worksheet = workbook[sheet_name]
                # Verify worksheet is readable
                data = [[cell.value for cell in row] for row in worksheet.iter_rows()]
                assert len(data) > 1  # Has headers + data
        except Exception as e:
            pytest.fail(f"Excel file validation failed: {str(e)}")
    
    def test_report_model_creation(self, excel_service):
        """Test Report model creation and validation"""
        report = Report(
            filename="test_report.xlsx",
            sheets=["ACQ", "Dials"],
            created_date="2025-07-28",
            total_records=100
        )
        
        assert report.filename == "test_report.xlsx"
        assert len(report.sheets) == 2
        assert "ACQ" in report.sheets
        assert report.total_records == 100
    
    @pytest.mark.performance
    def test_concurrent_excel_generation(self, excel_service, sample_report_data, temp_test_dir):
        """Test concurrent Excel file generation"""
        import threading
        import time
        
        def create_report(index):
            output_file = temp_test_dir / f"concurrent_report_{index}.xlsx"
            excel_service.create_excel_report(sample_report_data, output_file)
            return output_file.exists()
        
        # Create multiple reports concurrently
        threads = []
        results = []
        
        for i in range(3):
            thread = threading.Thread(target=lambda i=i: results.append(create_report(i)))
            threads.append(thread)
            thread.start()
        
        for thread in threads:
            thread.join()
        
        assert all(results)  # All reports created successfully
